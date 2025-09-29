# app/routers/reportes.py
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
import pandas as pd
from fastapi.responses import StreamingResponse
from app.database import SessionLocal
from app.schemas.lotes import LoteRead  
from app.crud.productos import obtener_productos_total# solo si usarás response_model
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from app.crud.lotes import obtener_lotes_por_rango
from app.crud.movimientos import obtener_movimientos
from app.crud.lotes import obtener_por_id
from app.crud.productos import obtener_por_id
from app.crud.reportes import contar_lotes_stock_bajo, generar_reporte_lotes, generar_reporte_movimientos


# Calcular fecha actual y fecha + 30 días para vencer
fecha_hoy = datetime.now().date()
fecha_en_30_dias = fecha_hoy + timedelta(days=30)

fecha_hoy, fecha_en_30_dias

router = APIRouter(prefix="/reportes", tags=["reportes"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/StockBajo")
def vista_previa_lotes(
    db: Session = Depends(get_db),
):
    datos = contar_lotes_stock_bajo(db)
    return  datos

@router.get("/lotes/vista")
def vista_previa_lotes(
    db: Session = Depends(get_db),
):
    lotes_bajo = contar_lotes_stock_bajo(db)
    set_bajo_ids = set(l.get("IdLote") for l in lotes_bajo.get("data", []) if "IdLote" in l)

    datos = generar_reporte_lotes(db)
    _, productos = obtener_productos_total(db)
    mapa_productos = {p.IdProducto: p for p in productos}

    resultados_debug = []
    for lote in datos:
        producto = mapa_productos.get(lote.IdProducto)

        resultados_debug.append({
            "IdLote": lote.IdLote,
            "NumeroLote": lote.NumeroLote,
            "Producto": producto.Nombre if producto else None,
            "CantidadActual": lote.CantidadActual,
            "StockMinimo": producto.StockMinimo if producto else None,
            "EnStockBajo": lote.IdLote in set_bajo_ids
        })

    return {
        "count": len(set_bajo_ids),
        "set_bajo": list(set_bajo_ids),
        "debug_data": resultados_debug
    }



@router.get("/lotes/excel")
def descargar_excel_lotes(
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    datos = generar_reporte_lotes(db, fecha_inicio, fecha_fin)
    if not datos:
        raise HTTPException(status_code=404, detail="No hay lotes para exportar.")

    _, productos = obtener_productos_total(db)
    mapa_productos = {p.IdProducto: p for p in productos}

    hoy = datetime.now().date()
    hora_actual = datetime.now().strftime("%H:%M:%S")
    vencimiento_limite = hoy + timedelta(days=30)
    total_vencer, _ = obtener_lotes_por_rango(
        db=db,
        skip=0,
        limit=10_000_000,
        vencimiento_inicio=hoy,
        vencimiento_fin=vencimiento_limite
    )
    lotes_bajo = contar_lotes_stock_bajo(db)
    total_bajo = lotes_bajo["count"]
    set_bajo = set(l["IdLote"] for l in lotes_bajo.get("data", []) if "IdLote" in l)

    # Estilos
    header_fill    = PatternFill("solid", fgColor="4F81BD")
    zebra_fill     = PatternFill("solid", fgColor="F2F2F2")
    fill_stock_low= PatternFill("solid", fgColor="FFCCCC")
    fill_vencer    = PatternFill("solid", fgColor="FFE699")
    border        = Border(*[Side(style="thin")] * 4)
    font_header   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    font_default  = Font(name="Calibri", size=11)
    font_bold     = Font(bold=True, name="Calibri", size=11)
    font_title    = Font(bold=True, name="Calibri", size=20)
    center_align  = Alignment("center", "center")
    right_align   = Alignment(horizontal="right")
    date_fmt      = "DD/MM/YYYY"

    wb = Workbook()
    ws = wb.active
    ws.title = "Lotes"

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = "REPORTE DE LOTES"
    ws["A1"].font = font_title
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 27
    ws.column_dimensions['A'].width = 6

    # Subtítulo Datos
    ws["C2"] = "Datos"
    ws["C2"].border = border
    ws["C2"].font = font_bold
    ws["C2"].alignment = center_align

    # Metadatos
    meta = [
        ("Cantidad Lotes:", len(datos)),
        ("Filtro Fecha Inicio:", fecha_inicio or "—"),
        ("Filtro Fecha Fin:", fecha_fin or "—"),
        ("Lotes Stock Bajo:", total_bajo),
        ("Lotes por Vencer (<30d):", total_vencer),
    ]
    for idx, (lbl, val) in enumerate(meta, start=3):
        ws[f"B{idx}"] = lbl
        ws[f"B{idx}"].font = font_bold
        ws[f"B{idx}"].alignment = Alignment(horizontal="left")
        ws[f"C{idx}"] = val
        ws[f"B{idx}"].border = ws[f"C{idx}"].border = border
    ws["B6"].fill = fill_stock_low
    ws["B7"].fill = fill_vencer

    # Fecha y hora
    ws["G3"] = "Fecha Creación"
    ws["G3"].border = border
    ws["G3"].font = font_bold
    ws["G3"].alignment = center_align
    ws["F4"] = "Fecha:"
    ws["F4"].font = font_bold
    ws["F4"].alignment = right_align
    ws["G4"] = hoy
    ws["G4"].number_format = date_fmt
    ws["G4"].alignment = right_align
    ws["F5"] = "Hora:"
    ws["F5"].font = font_bold
    ws["F5"].alignment = right_align
    ws["G5"] = hora_actual
    ws["G5"].alignment = right_align
    for r in (4,5): ws[f"F{r}"].border = ws[f"G{r}"].border = border

    # Encabezados tabla
    headers = ["N°","Código","Producto","Cantidad Inicial","Cantidad Actual","Fecha Fabricación","Fecha Recepción","Fecha Vencimiento"]
    ws.append([])
    ws.append(headers)
    ws.auto_filter.ref = "A9:H9"
    ws.freeze_panes = "A10"
    for cell in ws[9]:
        cell.fill = header_fill; cell.font = font_header; cell.alignment = center_align; cell.border = border

    # Datos y marcas
    low_rows = set(); venc_rows = set()
    for i, lote in enumerate(datos, start=1):
        ws.append([
            i,
            lote.NumeroLote or "—",
            None,
            lote.CantidadInicial or 0,
            lote.CantidadActual or 0,
            lote.FechaFabricacion,
            lote.FechaRecepcion,
            lote.FechaVencimiento,
        ])
        r = ws.max_row
        prod = mapa_productos.get(lote.IdProducto)
        ws.cell(r,3).value = prod.Nombre if prod else f"[{lote.IdProducto}]"
        if prod:
            ws.cell(r,3).comment = Comment(f"Código: {prod.CodigoProducto}\nUnidad: {prod.UnidadMedida}\nMínimo: {prod.StockMinimo} ","Sistema")
        if lote.IdLote in set_bajo: low_rows.add(r)
        if hoy <= lote.FechaVencimiento <= vencimiento_limite: venc_rows.add(r)

    # Zebra y borders
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border; cell.font = font_default
            if cell.row %2==0: cell.fill = zebra_fill

    # Destacar bajos y por vencer
    for r in low_rows:
        ws.cell(r,5).fill = fill_stock_low
    for r in venc_rows:
        ws.cell(r,8).fill = fill_vencer

    # Fechas
    for col in (6,7,8):
        for r in range(10, ws.max_row+1): ws.cell(r,col).number_format = date_fmt

    # Auto ancho
    for idx, col in enumerate(ws.columns, start=1):
        letra = get_column_letter(idx)
        if letra == 'A':
            continue  # No modificar ancho de columna A
        m = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(idx)].width = m+3
        
    
    

    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment; filename=reporte_lotes.xlsx"})

@router.get("/movimientos/excel")
def descargar_excel_movimientos(
    fecha_inicio: date = Query(None),
    fecha_fin: date = Query(None),
    db: Session = Depends(get_db),
):
    datos = generar_reporte_movimientos(db, fecha_inicio, fecha_fin)
    if not datos:
        raise HTTPException(status_code=404, detail="No hay movimientos para exportar.")

    hoy = datetime.now().date()
    hora_actual = datetime.now().strftime("%H:%M:%S")
    total_entrada = sum(1 for m in datos if m["TipoMovimiento"].lower() == "entrada")
    total_salida = sum(1 for m in datos if m["TipoMovimiento"].lower() == "salida")
    total_anulados = sum(1 for m in datos if str(m["FueAnulado"]).strip().upper() == "SI")

    # Estilos
    header_fill = PatternFill("solid", fgColor="4F81BD")
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    fill_anulado = PatternFill("solid", fgColor="FFCCCC")
    fill_entrada = PatternFill("solid", fgColor="C6EFCE")  # verde claro
    fill_salida = PatternFill("solid", fgColor="FFF2CC")   # amarillo claro
    border = Border(*[Side(style="thin")] * 4)
    font_header = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    font_default = Font(name="Calibri", size=11)
    font_bold = Font(bold=True, name="Calibri", size=11)
    font_title = Font(bold=True, name="Calibri", size=20)
    center_align = Alignment("center", "center")
    right_align = Alignment(horizontal="right")
    date_fmt = "DD/MM/YYYY"

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Título
    ws.merge_cells("A1:J1")
    ws["A1"] = "REPORTE DE MOVIMIENTOS"
    ws["A1"].font = font_title
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 27
    ws.column_dimensions['A'].width = 6  # ancho específico

    # Subtítulo Datos
    ws["C2"] = "Datos"
    ws["C2"].border = border
    ws["C2"].font = font_bold
    ws["C2"].alignment = center_align

    # Metadatos
    meta = [
        ("Cantidad de Movimientos de Entrada:", total_entrada, fill_entrada),
        ("Cantidad de Movimientos de Salida:", total_salida, fill_salida),
        ("Movimientos Anulados:", total_anulados, fill_anulado),
        ("Filtro Fecha Inicio:", fecha_inicio or "—", None),
        ("Filtro Fecha Fin:", fecha_fin or "—", None),
    ]
    for idx, (lbl, val, color) in enumerate(meta, start=3):
        ws[f"B{idx}"] = lbl
        ws[f"C{idx}"] = val
        ws[f"B{idx}"].font = font_bold
        ws[f"C{idx}"].font = font_default
        ws[f"B{idx}"].alignment = Alignment(horizontal="left")
        ws[f"B{idx}"].border = ws[f"C{idx}"].border = border
        if color:
            ws[f"B{idx}"].fill = color

    # Fecha y hora
    ws["H3"] = "Fecha Creación"
    ws["H3"].border = border
    ws["H3"].font = font_bold
    ws["H3"].alignment = center_align
    ws["G4"] = "Fecha:"
    ws["G4"].font = font_bold
    ws["G4"].alignment = right_align
    ws["H4"] = hoy
    ws["H4"].number_format = date_fmt
    ws["H4"].alignment = right_align
    ws["G5"] = "Hora:"
    ws["G5"].font = font_bold
    ws["G5"].alignment = right_align
    ws["H5"] = hora_actual
    ws["H5"].alignment = right_align
    for r in (4, 5):
        ws[f"G{r}"].border = ws[f"H{r}"].border = border

    # Encabezados con nuevo orden
    headers = [
        "N°", "Lote","Tipo", "Cantidad", "Anulado",
        "Motivo", "Responsable", "Documento", "Observaciones", "Fecha"
    ]
    ws.append([])
    ws.append(headers)
    ws.auto_filter.ref = "A9:J9"
    ws.freeze_panes = "A10"
    for cell in ws[9]:
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = center_align
        cell.border = border

    # Filas de datos
    for i, m in enumerate(datos, start=1):
        fue_anulado = str(m["FueAnulado"]).strip().upper() == "SI"
        tipo = m["TipoMovimiento"].lower()
        fila = [
            i,
            m["NombreLote"],
            m["TipoMovimiento"],
            m["Cantidad"],
            "Sí" if fue_anulado else "No",
            m["Motivo"],
            m["Responsable"],
            m["DocumentoReferencia"],
            m["Observaciones"],
            m["FechaMovimiento"],
        ]
        ws.append(fila)
        r = ws.max_row

        # Comentario
        if m.get("ComentarioLote"):
            ws.cell(r, 2).comment = Comment(m["ComentarioLote"], "Sistema")

        # Zebra
        for col_idx, cell in enumerate(ws[r], start=1):
            cell.border = border
            cell.font = font_default
            if r % 2 == 0:
                cell.fill = zebra_fill

        # Pintar tipo
        tipo_cell = ws.cell(r, 3)
        if tipo == "entrada":
            tipo_cell.fill = fill_entrada
        elif tipo == "salida":
            tipo_cell.fill = fill_salida

        # Pintar anulado
        if fue_anulado:
            ws.cell(r, 5).fill = fill_anulado

    # Formato fecha columna "J"
    for r in range(10, ws.max_row + 1):
        ws.cell(r, 10).number_format = date_fmt

    # Ajuste automático
    for idx, col in enumerate(ws.columns, start=1):
        letra = get_column_letter(idx)
        if letra == 'A':
            continue
        m = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(idx)].width = max(m + 3, 10)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_movimientos.xlsx"}
    )
